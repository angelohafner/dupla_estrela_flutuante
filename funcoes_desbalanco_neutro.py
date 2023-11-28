from io import BytesIO
import matplotlib.pyplot as plt
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Color
import colorsys
import copy
import psutil


# %%


# %%
def exporta_para_excel(planilha='Ia1', arquivo='correntes.xlsx', valor_nominal=10.1):
    wb = load_workbook(arquivo)
    ws = wb[planilha]
    # Carregar a pasta de trabalho de origem e a planilha
    src_wb = load_workbook('matriz_total.xlsx')
    src_ws = src_wb['A']

    # Carregar a pasta de trabalho de destino e a planilha
    dest_wb = load_workbook(arquivo)
    dest_ws = dest_wb[planilha]

    # Iterar sobre todas as células na planilha de origem
    for row in src_ws.iter_rows():
        for cell in row:
            # Obter a célula correspondente na planilha de destino
            dest_cell = dest_ws[cell.coordinate]

            # Copiar a formatação da célula de origem para a célula de destino
            dest_cell.font = copy.copy(cell.font)
            dest_cell.border = copy.copy(cell.border)
            dest_cell.fill = copy.copy(cell.fill)
            dest_cell.number_format = cell.number_format
            dest_cell.protection = copy.copy(cell.protection)
            dest_cell.alignment = copy.copy(cell.alignment)
            dest_cell.number_format = "##0.0E+00"

    # Salvar a pasta de trabalho de destino
    dest_wb.save(arquivo)


# %%
def matriz_complexa_para_polar(matriz_complexa):
    magnitude = np.abs(matriz_complexa)
    angulo = np.angle(matriz_complexa)

    # Gera a matriz de strings em formato polar
    matriz_polar = np.empty(matriz_complexa.shape, dtype=object)

    for i in range(matriz_complexa.shape[0]):
        for j in range(matriz_complexa.shape[1]):
            matriz_polar[i, j] = f'{magnitude[i, j]:.2f}∠{np.degrees(angulo[i, j]):.2f}°'

    return matriz_polar


# %%
def load_excel_to_numpy(filename):
    wb = load_workbook(filename=filename, read_only=True)
    sheets = ['A', 'B', 'C']
    data = {}

    for sheet in sheets:
        ws = wb[sheet]
        data[sheet] = pd.DataFrame(ws.values)

    return {k: v.to_numpy() for k, v in data.items()}


# %%
def matriz_fases_ramos(nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int, cap_interna, des_int):
    matriz = np.ones((3, nr_lin_ext * nr_lin_int, 2 * nr_col_ext * nr_col_int))
    cores = ['FFFFCC', 'CCFFCC', 'CCCCFF', 'FFCCFF']

    wb = Workbook()
    ws = wb.active
    ws.title = "A"

    for fase in ['A', 'B', 'C']:
        if fase == 'A':
            ws = wb.active
            fase_cont = 0
        else:
            ws = wb.create_sheet(title=fase)
            fase_cont = fase_cont + 1

        for i_ext in range(nr_lin_ext):
            for j_ext in range(2 * nr_col_ext):
                cor = cores[(i_ext + j_ext) % len(cores)]
                fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
                for i_int in range(nr_lin_int):
                    for j_int in range(nr_col_int):
                        ii = i_ext * nr_lin_int + i_int
                        jj = j_ext * nr_col_int + j_int
                        valor = cap_interna * np.random.uniform(1 - des_int, 1 + des_int)
                        matriz[fase_cont, ii, jj] = valor
                        ws.cell(row=ii + 1, column=jj + 1, value=valor)
                        ws.cell(row=ii + 1, column=jj + 1).fill = fill
                        ws.cell(row=ii + 1, column=jj + 1).number_format = '##0.0E+#0'
                        ws.column_dimensions[get_column_letter(jj + 1)].width = 8
                        ws.row_dimensions[ii + 1].height = 30
                        ws.cell(row=ii + 1, column=jj + 1).alignment = Alignment(horizontal='center', vertical='center')
                        if j_ext >= nr_col_ext:
                            ws.cell(row=ii + 1, column=jj + 1).font = Font(name='Bahnschrift SemiBold Condensed',
                                                                           color="FF0000", underline="single")
                        else:
                            ws.cell(row=ii + 1, column=jj + 1).font = Font(name='Bahnschrift SemiBold Condensed',
                                                                           color="0000FF")

    return wb, matriz


# %%
def matrizes_internas(matriz_FR, nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int):
    super_matriz_FR = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int))
    eq_paral_internos = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int))
    eq_serie_internos = np.ones((nr_lin_ext, nr_col_ext))
    for i_ext in range(nr_lin_ext):
        for j_ext in range(nr_col_ext):
            rs = i_ext * nr_lin_int
            re = rs + nr_lin_int
            cs = j_ext * nr_col_int
            ce = cs + nr_col_int
            super_matriz_FR[i_ext, j_ext, :, :] = np.array(matriz_FR[rs:re, cs:ce])
            sbmatriz = super_matriz_FR[i_ext, j_ext, :, :]
            eq_paral_internos[i_ext, j_ext, :] = np.sum(sbmatriz, axis=1)
            eq_serie_internos[i_ext, j_ext] = 1 / np.sum(1 / eq_paral_internos[i_ext, j_ext, :])

    return super_matriz_FR, eq_paral_internos, eq_serie_internos


# %%
def calcular_corrente_tensao(V_ao, omega, eq_serie_externos_A1, eq_paral_externos_A1, eq_serie_internos_A1,
                             eq_paral_internos_A1, matriz_A1):
    I_a1 = V_ao * (1j * omega * eq_serie_externos_A1)
    V_a1_ser = I_a1 * 1 / (1j * omega * eq_paral_externos_A1)
    # V_a1_ser = V_a1_ser.reshape(-1, 1) # necessário para transformar em matriz com uma coluna
    I_a1_par = V_a1_ser * (1j * omega * eq_serie_internos_A1)
    V_a1_ser_int = I_a1_par * 1 / (1j * omega * eq_paral_internos_A1)
    I_a1_par_int = V_a1_ser_int * (1j * omega * matriz_A1)
    return I_a1, V_a1_ser, I_a1_par, V_a1_ser_int, I_a1_par_int


# %%
def colorir_planilha_especifica(wb, planilha, nr_lin_int, nr_col_int, nr_lin_ext, nr_col_ext):
    cores = ['FFFFCC', 'CCFFCC', 'CCCCFF', 'FFCCFF']
    ws = wb[planilha]
    for i_ext in range(nr_lin_ext):
        for j_ext in range(2 * nr_col_ext):
            cor = cores[(i_ext + j_ext) % len(cores)]
            fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
            for i_int in range(nr_lin_int):
                for j_int in range(nr_col_int):
                    ii = i_ext * nr_lin_int + i_int
                    jj = j_ext * nr_col_int + j_int
                    ws.row_dimensions[ii + 1].height = 30
                    ws.cell(row=ii + 1, column=jj + 1).alignment = Alignment(horizontal='center', vertical='center')
                    if j_ext >= nr_col_ext:
                        ws.cell(row=ii + 1, column=jj + 1).font = Font(name='Bahnschrift SemiBold Condensed',
                                                                       color="FF0000", underline="single")
                    else:
                        ws.cell(row=ii + 1, column=jj + 1).font = Font(name='Bahnschrift SemiBold Condensed',
                                                                       color="0000FF")

    wb.save('correntes.xlsx')


# %%
def fase(matriz, nr_col_ext, nr_col_int, nr_lin_ext, nr_lin_int, ramo1, ramo2):
    matriz1 = matriz[:, :nr_col_ext * nr_col_int]  # vai até a metade da matriz
    matriz2 = matriz[:, nr_col_ext * nr_col_int:]  # vem a partir da metade da matriz

    super_matriz1, eq_paral_internos1, eq_serie_internos1 = matrizes_internas(matriz_FR=matriz1, nr_lin_ext=nr_lin_ext,
                                                                              nr_col_ext=nr_col_ext,
                                                                              nr_lin_int=nr_lin_int,
                                                                              nr_col_int=nr_col_int)
    super_matriz2, eq_paral_internos2, eq_serie_internos2 = matrizes_internas(matriz_FR=matriz2, nr_lin_ext=nr_lin_ext,
                                                                              nr_col_ext=nr_col_ext,
                                                                              nr_lin_int=nr_lin_int,
                                                                              nr_col_int=nr_col_int)

    eq_paral_externos1 = (np.sum(eq_serie_internos1, axis=1)).reshape(-1, 1)
    eq_serie_externos1 = 1 / np.sum(1 / eq_paral_externos1)

    eq_paral_externos2 = (np.sum(eq_serie_internos2, axis=1)).reshape(-1, 1)
    eq_serie_externos2 = 1 / np.sum(1 / eq_paral_externos2)

    eq_paral_ramos = eq_serie_externos1 + eq_serie_externos2

    return [eq_paral_ramos, eq_serie_externos1, eq_paral_externos1, eq_serie_internos1, eq_paral_internos1,
            eq_serie_externos2, eq_paral_externos2, eq_serie_internos2, eq_paral_internos2, matriz1, matriz2,
            super_matriz1, super_matriz2]


# %%
def calcular_correntes_tensoes(Za, Zb, Zc, Vff, a, matriz_impedancia_sistema):
    matriz_impedancia_malha = np.array([[Za + Zb, -Zb], [-Zb, Zb + Zc]])
    matriz_fontes_malha = np.array([[Vff], [Vff * (a ** 2)]])
    matriz_correntes_alphabeta = np.linalg.inv(matriz_impedancia_malha) @ matriz_fontes_malha
    I_alpha = matriz_correntes_alphabeta[0, 0]
    I_beta = matriz_correntes_alphabeta[1, 0]
    matriz_correntes_fase = np.array([[I_alpha], [I_beta - I_alpha], [-I_beta]])
    matriz_tensoes_Vabco = matriz_impedancia_sistema @ matriz_correntes_fase
    tensao_deslocamento_netro = np.sum(matriz_tensoes_Vabco) / 3
    return matriz_correntes_fase, matriz_tensoes_Vabco, tensao_deslocamento_netro


def salvar_dataframes_com_exporta(I_c1, I_c2, arquivo='correntes.xlsx', planilha_c1='externos-c1',
                                  planilha_c2='externos-c2', planilha_c='externos-c'):
    df_I_c1_par_ext_abs = pd.DataFrame(np.abs(I_c1))
    df_I_c2_par_ext_abs = pd.DataFrame(np.abs(I_c2))
    df_I_c_par_ext_abs = pd.concat([df_I_c1_par_ext_abs, df_I_c2_par_ext_abs], axis=1)
    df_I_c1_par_ext_arg = pd.DataFrame(np.angle(I_c1))
    df_I_c2_par_ext_arg = pd.DataFrame(np.angle(I_c2))
    df_I_c_par_ext_arg = pd.concat([df_I_c1_par_ext_arg, df_I_c2_par_ext_arg], axis=1)

    with pd.ExcelWriter(arquivo, engine='openpyxl', mode='a') as writer:
        df_I_c_par_ext_abs.to_excel(writer, sheet_name=planilha_c+'_abs', index=False, header=False)
        df_I_c_par_ext_arg.to_excel(writer, sheet_name=planilha_c+'_arg', index=False, header=False)

    if planilha_c.find("internos") != -1:
        exporta_para_excel(planilha=planilha_c+'_abs', arquivo=arquivo)
        exporta_para_excel(planilha=planilha_c+'_arg', arquivo=arquivo)


# def salvar_dataframes_com_exporta_A1(I_c1, I_c2, arquivo='correntes.xlsx', planilha_c1='externos-c1',
#                                   planilha_c2='externos-c2', planilha_c='externos-c'):
#
#     df_I_c1_par_ext_abs = pd.DataFrame(np.abs(I_c1))
#     df_I_c2_par_ext_abs = pd.DataFrame(np.abs(I_c2))
#     df_I_c_par_ext_abs = pd.concat([df_I_c1_par_ext_abs, df_I_c2_par_ext_abs], axis=1)
#
#     with pd.ExcelWriter(arquivo, engine='openpyxl', mode='w') as writer:
#         # df_I_c1_par_ext_abs.to_excel(writer, sheet_name=planilha_c1, index=False, header=False)
#         # df_I_c2_par_ext_abs_abs.to_excel(writer, sheet_name=planilha_c2, index=False, header=False)
#         df_I_c_par_ext_abs.to_excel(writer, sheet_name=planilha_c, index=False, header=False)
#
#     # exporta_para_excel(planilha=planilha_c1, arquivo=arquivo)
#     # exporta_para_excel(planilha=planilha_c2, arquivo=arquivo)
#     if planilha_c.find("interno") != -1:
#         exporta_para_excel(planilha=planilha_c, arquivo=arquivo)


def calculate_matrices(nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int, I_c2_par_ext, omega, eq_paral_internos_C2,
                       super_matriz_C2):
    I_c2_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
    V_c2_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int), dtype='complex')
    V_c2_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')

    for i_ext in range(nr_lin_ext):
        for j_ext in range(nr_col_ext):
            V_c2_ser_int[i_ext, j_ext, :] = I_c2_par_ext[i_ext, j_ext] / (
                    1j * omega * eq_paral_internos_C2[i_ext, j_ext, :])
            I_c2_par_int[i_ext, j_ext, :, :] = V_c2_ser_int[i_ext, j_ext, :].reshape(-1, 1) * (
                    1j * super_matriz_C2[i_ext, j_ext, :, :])
            V_c2_par_int[i_ext, j_ext, :, :] = I_c2_par_int[i_ext, j_ext, :, :] / (
                    1j * super_matriz_C2[i_ext, j_ext, :, :])

    I_c2_par_int_excel = np.ones((nr_lin_ext * nr_lin_int, nr_col_ext * nr_col_int), dtype='complex')
    V_c2_par_int_excel = np.ones((nr_lin_ext * nr_lin_int, nr_col_ext * nr_col_int), dtype='complex')

    for i_ext in range(nr_lin_ext):
        for j_ext in range(nr_col_ext):
            rs = i_ext * nr_lin_int
            re = rs + nr_lin_int
            cs = j_ext * nr_col_int
            ce = cs + nr_col_int
            I_c2_par_int_excel[rs:re, cs:ce] = I_c2_par_int[i_ext, j_ext, :, :]
            V_c2_par_int_excel[rs:re, cs:ce] = V_c2_par_int[i_ext, j_ext, :, :]

    return I_c2_par_int, V_c2_ser_int, V_c2_par_int, I_c2_par_int_excel, V_c2_par_int_excel


# You can call this function and pass the required arguments to get the desired matrices.


import numpy as np


def compute_values(V_ao, omega,
                   eq_serie_externos_C2, eq_paral_externos_C2,
                   eq_serie_internos_C2, eq_paral_internos_C2,
                   nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int,
                   super_matriz_C2):
    I_c2_ext = V_ao * (1j * omega * eq_serie_externos_C2)
    V_c2_ser_ext = I_c2_ext / (1j * omega * eq_paral_externos_C2)
    I_c2_par_ext = V_c2_ser_ext * (1j * omega * eq_serie_internos_C2)

    I_c2_par_int, V_c2_ser_int, V_c2_par_int, I_c2_par_int_excel, V_c2_par_int_excel = calculate_matrices(
        nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int, I_c2_par_ext, omega, eq_paral_internos_C2, super_matriz_C2
    )

    return I_c2_ext, V_c2_ser_ext, I_c2_par_ext, I_c2_par_int, V_c2_ser_int, V_c2_par_int, I_c2_par_int_excel, V_c2_par_int_excel


# Assuming the function `calculate_matrices` is defined somewhere before using `compute_values`


import os
from openpyxl import Workbook, load_workbook
import pandas as pd


def save_data_to_excel(eq_data, filename="filename", sheet_name="sheet_name",
                       nr_lin_ext=None, nr_col_ext=None, nr_lin_int=None, nr_col_int=None):
    """
    Save a 3D numpy array or list to an Excel file.

    Parameters:
    - eq_data: 3D numpy array or list to save.
    - filename: Name of the Excel file.
    - sheet_name: Name of the Excel sheet.
    - nr_lin_ext: Number of outer rows in eq_data.
    - nr_col_ext: Number of outer columns in eq_data.
    - nr_lin_int: Number of inner rows in eq_data.
    - nr_col_int: Number of inner columns in eq_data.
    """

    # Load the workbook if exists, otherwise create a new one
    if os.path.exists(filename):
        book = load_workbook(filename)
    else:
        book = Workbook()

    # Check if the sheet exists, if not create it
    if sheet_name not in book.sheetnames:
        ws = book.create_sheet(sheet_name)
    else:
        ws = book[sheet_name]

    # Write to the workbook using openpyxl directly
    for i in range(nr_lin_ext):
        for j in range(nr_col_ext):
            dataframe = pd.DataFrame(eq_data[i, j, :])

            # Calculate starting positions
            start_row = i * nr_lin_int
            start_col = j * nr_col_int

            # Write to the sheet
            for df_row, row in enumerate(dataframe.values, start=start_row + 1):  # +1 because openpyxl is 1-based index
                for df_col, value in enumerate(row, start=start_col + 1):
                    ws.cell(row=df_row, column=df_col, value=value)

    # Save the modified workbook
    book.save(filename)


def save_multiple_datasets_to_excel(datasets, filename, sheet_names, nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int):
    for data, sheet_name in zip(datasets, sheet_names):
        save_data_to_excel(data, filename, sheet_name, nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int)


def append_df_to_excel(matriz, filename="matriz_total.xlsx", sheet_name="eq_serie_internos_A1"):
    df = pd.DataFrame(matriz)
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
