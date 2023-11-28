import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule
from engineering_notation import EngNumber
from funcoes_desbalanco_neutro import *


# %% Fase A, Ramo 1
# corrente do capacitor equivalente
I_a1_ext = V_ao * (1j*omega*eq_serie_externos_A1)
# tensoes dos capacitores que estao em serie
V_a1_ser_ext = I_a1_ext / (1j*omega*eq_paral_externos_A1)
# corrente dos capacitores que estao em paralelo
I_a1_par_ext = V_a1_ser_ext * (1j*omega*eq_serie_internos_A1)

# AGORA VAMOS PARA OS INTERNOS
I_a1_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
V_a1_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int), dtype='complex')
V_a1_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        # capacitores internos em série
        V_a1_ser_int[i_ext, j_ext, :] = I_a1_par_ext[i_ext, j_ext] / (1j * omega * eq_paral_internos_A1[i_ext, j_ext, :])
        I_a1_par_int[i_ext, j_ext, :, :] = V_a1_ser_int[i_ext, j_ext, :].reshape(-1, 1) * (1j*super_matriz_A1[i_ext, j_ext, :, :])
        V_a1_par_int[i_ext, j_ext, :, :] = I_a1_par_int[i_ext, j_ext, :, :] / (1j * super_matriz_A1[i_ext, j_ext, :, :])

I_a1_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
V_a1_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        rs = i_ext * nr_lin_int
        re = rs + nr_lin_int
        cs = j_ext * nr_col_int
        ce = cs + nr_col_int
        I_a1_par_int_excel[rs:re, cs:ce] = I_a1_par_int[i_ext, j_ext, :, :]
        V_a1_par_int_excel[rs:re, cs:ce] = V_a1_par_int[i_ext, j_ext, :, :]



df_I_a1_par_int_excel = pd.DataFrame(np.abs(I_a1_par_int_excel))
filename = 'correntes.xlsx'
sheetname = 'internos-a1'
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df_I_a1_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_V_a1_par_int_excel = pd.DataFrame(np.abs(V_a1_par_int_excel))
filename = 'tensoes.xlsx'
sheetname = 'internos-a1'
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df_V_a1_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_I_a1_par_ext = pd.DataFrame(np.abs(I_a1_par_ext))
filename = 'correntes.xlsx'
sheetname = 'externos-a1'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_a1_par_ext.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)



# %% Fase A, Ramo 2
# corrente do capacitor equivalente
I_a2_ext = V_ao * (1j*omega*eq_serie_externos_A2)
# tensoes dos capacitores que estao em serie
V_a2_ser_ext = I_a2_ext / (1j*omega*eq_paral_externos_A2)
# corrente dos capacitores que estao em paralelo
I_a2_par_ext = V_a2_ser_ext * (1j*omega*eq_serie_internos_A2)

# AGORA VAMOS PARA OS INTERNOS
I_a2_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
V_a2_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int), dtype='complex')
V_a2_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        # capacitores internos em série
        V_a2_ser_int[i_ext, j_ext, :] = I_a2_par_ext[i_ext, j_ext] / (1j * omega * eq_paral_internos_A2[i_ext, j_ext, :])
        I_a2_par_int[i_ext, j_ext, :, :] = V_a2_ser_int[i_ext, j_ext, :].reshape(-1, 1) * (1j*super_matriz_A2[i_ext, j_ext, :, :])
        V_a2_par_int[i_ext, j_ext, :, :] = I_a2_par_int[i_ext, j_ext, :, :] / (1j * super_matriz_A2[i_ext, j_ext, :, :])

I_a2_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
V_a2_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        rs = i_ext * nr_lin_int
        re = rs + nr_lin_int
        cs = j_ext * nr_col_int
        ce = cs + nr_col_int
        I_a2_par_int_excel[rs:re, cs:ce] = I_a2_par_int[i_ext, j_ext, :, :]
        V_a2_par_int_excel[rs:re, cs:ce] = V_a2_par_int[i_ext, j_ext, :, :]



df_I_a2_par_int_excel = pd.DataFrame(np.abs(I_a2_par_int_excel))
filename = 'correntes.xlsx'
sheetname = 'internos-a2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_a2_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_V_a2_par_int_excel = pd.DataFrame(np.abs(V_a2_par_int_excel))
filename = 'tensoes.xlsx'
sheetname = 'internos-a2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_V_a2_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_I_a2_par_ext = pd.DataFrame(np.abs(I_a2_par_ext))
filename = 'correntes.xlsx'
sheetname = 'externos-a2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_a2_par_ext.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

# %% Fase B, Ramo 1
# corrente do capacitor equivalente
I_b1_ext = V_ao * (1j*omega*eq_serie_externos_B1)
# tensoes dos capacitores que estao em serie
V_b1_ser_ext = I_b1_ext / (1j*omega*eq_paral_externos_B1)
# corrente dos capacitores que estao em paralelo
I_b1_par_ext = V_b1_ser_ext * (1j*omega*eq_serie_internos_B1)

# AGORA VAMOS PARA OS INTERNOS
I_b1_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
V_b1_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int), dtype='complex')
V_b1_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        # capacitores internos em série
        V_b1_ser_int[i_ext, j_ext, :] = I_b1_par_ext[i_ext, j_ext] / (1j * omega * eq_paral_internos_B1[i_ext, j_ext, :])
        I_b1_par_int[i_ext, j_ext, :, :] = V_b1_ser_int[i_ext, j_ext, :].reshape(-1, 1) * (1j*super_matriz_B1[i_ext, j_ext, :, :])
        V_b1_par_int[i_ext, j_ext, :, :] = I_b1_par_int[i_ext, j_ext, :, :] / (1j * super_matriz_B1[i_ext, j_ext, :, :])

I_b1_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
V_b1_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        rs = i_ext * nr_lin_int
        re = rs + nr_lin_int
        cs = j_ext * nr_col_int
        ce = cs + nr_col_int
        I_b1_par_int_excel[rs:re, cs:ce] = I_b1_par_int[i_ext, j_ext, :, :]
        V_b1_par_int_excel[rs:re, cs:ce] = V_b1_par_int[i_ext, j_ext, :, :]



df_I_b1_par_int_excel = pd.DataFrame(np.abs(I_b1_par_int_excel))
filename = 'correntes.xlsx'
sheetname = 'internos-b1'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_b1_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_V_b1_par_int_excel = pd.DataFrame(np.abs(V_b1_par_int_excel))
filename = 'tensoes.xlsx'
sheetname = 'internos-b1'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_V_b1_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_I_b1_par_ext = pd.DataFrame(np.abs(I_b1_par_ext))
filename = 'correntes.xlsx'
sheetname = 'externos-b1'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_b1_par_ext.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)


# %% Fase B, Ramo 2
# corrente do capacitor equivalente
I_b2_ext = V_ao * (1j*omega*eq_serie_externos_B2)
# tensoes dos capacitores que estao em serie
V_b2_ser_ext = I_b2_ext / (1j*omega*eq_paral_externos_B2)
# corrente dos capacitores que estao em paralelo
I_b2_par_ext = V_b2_ser_ext * (1j*omega*eq_serie_internos_B2)

# AGORA VAMOS PARA OS INTERNOS
I_b2_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
V_b2_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int), dtype='complex')
V_b2_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        # capacitores internos em série
        V_b2_ser_int[i_ext, j_ext, :] = I_b2_par_ext[i_ext, j_ext] / (1j * omega * eq_paral_internos_B2[i_ext, j_ext, :])
        I_b2_par_int[i_ext, j_ext, :, :] = V_b2_ser_int[i_ext, j_ext, :].reshape(-1, 1) * (1j*super_matriz_B2[i_ext, j_ext, :, :])
        V_b2_par_int[i_ext, j_ext, :, :] = I_b2_par_int[i_ext, j_ext, :, :] / (1j * super_matriz_B2[i_ext, j_ext, :, :])

I_b2_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
V_b2_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        rs = i_ext * nr_lin_int
        re = rs + nr_lin_int
        cs = j_ext * nr_col_int
        ce = cs + nr_col_int
        I_b2_par_int_excel[rs:re, cs:ce] = I_b2_par_int[i_ext, j_ext, :, :]
        V_b2_par_int_excel[rs:re, cs:ce] = V_b2_par_int[i_ext, j_ext, :, :]



df_I_b2_par_int_excel = pd.DataFrame(np.abs(I_b2_par_int_excel))
filename = 'correntes.xlsx'
sheetname = 'internos-b2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_b2_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_V_b2_par_int_excel = pd.DataFrame(np.abs(V_b2_par_int_excel))
filename = 'tensoes.xlsx'
sheetname = 'internos-b2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_V_b2_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_I_b2_par_ext = pd.DataFrame(np.abs(I_b2_par_ext))
filename = 'correntes.xlsx'
sheetname = 'externos-b2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_b2_par_ext.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)


# %% Fase C, Ramo 1
# corrente do capacitor equivalente
I_c1_ext = V_ao * (1j*omega*eq_serie_externos_C1)
# tensoes dos capacitores que estao em serie
V_c1_ser_ext = I_c1_ext / (1j*omega*eq_paral_externos_C1)
# corrente dos capacitores que estao em paralelo
I_c1_par_ext = V_c1_ser_ext * (1j*omega*eq_serie_internos_C1)

# AGORA VAMOS PARA OS INTERNOS
I_c1_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
V_c1_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int), dtype='complex')
V_c1_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        # capacitores internos em série
        V_c1_ser_int[i_ext, j_ext, :] = I_c1_par_ext[i_ext, j_ext] / (1j * omega * eq_paral_internos_C1[i_ext, j_ext, :])
        I_c1_par_int[i_ext, j_ext, :, :] = V_c1_ser_int[i_ext, j_ext, :].reshape(-1, 1) * (1j*super_matriz_C1[i_ext, j_ext, :, :])
        V_c1_par_int[i_ext, j_ext, :, :] = I_c1_par_int[i_ext, j_ext, :, :] / (1j * super_matriz_C1[i_ext, j_ext, :, :])

I_c1_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
V_c1_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        rs = i_ext * nr_lin_int
        re = rs + nr_lin_int
        cs = j_ext * nr_col_int
        ce = cs + nr_col_int
        I_c1_par_int_excel[rs:re, cs:ce] = I_c1_par_int[i_ext, j_ext, :, :]
        V_c1_par_int_excel[rs:re, cs:ce] = V_c1_par_int[i_ext, j_ext, :, :]



df_I_c1_par_int_excel = pd.DataFrame(np.abs(I_c1_par_int_excel))
filename = 'correntes.xlsx'
sheetname = 'internos-c1'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_c1_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_V_c1_par_int_excel = pd.DataFrame(np.abs(V_c1_par_int_excel))
filename = 'tensoes.xlsx'
sheetname = 'internos-c1'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_V_c1_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_I_c1_par_ext = pd.DataFrame(np.abs(I_c1_par_ext))
filename = 'correntes.xlsx'
sheetname = 'externos-c1'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_c1_par_ext.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)


# %% Fase C, Ramo 2
# corrente do capacitor equivalente
I_c2_ext = V_ao * (1j*omega*eq_serie_externos_C2)
# tensoes dos capacitores que estao em serie
V_c2_ser_ext = I_c2_ext / (1j*omega*eq_paral_externos_C2)
# corrente dos capacitores que estao em paralelo
I_c2_par_ext = V_c2_ser_ext * (1j*omega*eq_serie_internos_C2)

# AGORA VAMOS PARA OS INTERNOS
I_c2_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
V_c2_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int), dtype='complex')
V_c2_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        # capacitores internos em série
        V_c2_ser_int[i_ext, j_ext, :] = I_c2_par_ext[i_ext, j_ext] / (1j * omega * eq_paral_internos_C2[i_ext, j_ext, :])
        I_c2_par_int[i_ext, j_ext, :, :] = V_c2_ser_int[i_ext, j_ext, :].reshape(-1, 1) * (1j*super_matriz_C2[i_ext, j_ext, :, :])
        V_c2_par_int[i_ext, j_ext, :, :] = I_c2_par_int[i_ext, j_ext, :, :] / (1j * super_matriz_C2[i_ext, j_ext, :, :])

I_c2_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
V_c2_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        rs = i_ext * nr_lin_int
        re = rs + nr_lin_int
        cs = j_ext * nr_col_int
        ce = cs + nr_col_int
        I_c2_par_int_excel[rs:re, cs:ce] = I_c2_par_int[i_ext, j_ext, :, :]
        V_c2_par_int_excel[rs:re, cs:ce] = V_c2_par_int[i_ext, j_ext, :, :]



df_I_c2_par_int_excel = pd.DataFrame(np.abs(I_c2_par_int_excel))
filename = 'correntes.xlsx'
sheetname = 'internos-c2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_c2_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_V_c2_par_int_excel = pd.DataFrame(np.abs(V_c2_par_int_excel))
filename = 'tensoes.xlsx'
sheetname = 'internos-c2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_V_c2_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_I_c2_par_ext = pd.DataFrame(np.abs(I_c2_par_ext))
filename = 'correntes.xlsx'
sheetname = 'externos-c2'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_c2_par_ext.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)