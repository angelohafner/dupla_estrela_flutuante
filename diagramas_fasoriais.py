from io import BytesIO
import matplotlib.pyplot as plt
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import cmath


def save_fasorial_image(V_ao, V_bo, V_co, V_on,
                        string_Vabco=['V_ao', 'V_bo', 'V_co', 'V_on'],
                        cores=['red', 'brown', 'yellow', 'blue'],
                        img_filename='Fasorial_Tensoes.png',
                        excel_filename='tensoes.xlsx',
                        sheet_name='Vabco'):

    plt.figure(figsize=(5, 5))
    plt.axvline(x=0, color='k', linestyle='--')
    plt.axhline(y=0, color='k', linestyle='--')
    tensoes = [V_ao, V_bo, V_co, V_on]
    labels = string_Vabco
    colors = cores

    for tensao, label, color in zip(tensoes, labels, colors):
        plt.quiver(0, 0, tensao.real, tensao.imag, angles='xy', scale_units='xy', scale=1, label=label, color=color)

    max_val = max([abs(val) for val in tensoes]) + 1
    plt.xlim(-max_val, max_val)
    plt.ylim(-max_val, max_val)
    plt.xlabel('Real')
    plt.ylabel('Imaginário')
    plt.grid(True)
    plt.legend()
    image_stream = BytesIO()
    plt.savefig(image_stream, format='png')
    image_stream.seek(0)  # Retorne para o início do buffer
    img = Image(image_stream)
    if not os.path.exists(excel_filename):
        wb = Workbook()
    else:
        wb = load_workbook(excel_filename)
    ws = wb.create_sheet(sheet_name)
    ws.add_image(img, 'A8')
    if sheet_name=='Vabco':
        V_ab = V_ao - V_bo
        V_bc = V_bo - V_co
        V_ca = V_co - V_ao
        data_dict = {
            'Vao': [abs(V_ao), cmath.phase(V_ao), V_ao.real, V_ao.imag, V_ao],
            'Vbo': [abs(V_bo), cmath.phase(V_bo), V_bo.real, V_bo.imag, V_bo],
            'Vco': [abs(V_co), cmath.phase(V_co), V_co.real, V_co.imag, V_co],
            'Von': [abs(V_on), cmath.phase(V_on), V_on.real, V_on.imag, V_on],
            'Vab': [abs(V_ab), cmath.phase(V_ab), V_ab.real, V_ab.imag, V_ab],
            'Vbc': [abs(V_bc), cmath.phase(V_bc), V_bc.real, V_bc.imag, V_bc],
            'Vca': [abs(V_ca), cmath.phase(V_ca), V_ca.real, V_ca.imag, V_ca],
        }
    else:
        data_dict = {
            'Iao': [abs(V_ao), cmath.phase(V_ao), V_ao.real, V_ao.imag, V_ao],
            'Ibo': [abs(V_bo), cmath.phase(V_bo), V_bo.real, V_bo.imag, V_bo],
            'Ico': [abs(V_co), cmath.phase(V_co), V_co.real, V_co.imag, V_co],
            'Ion': [abs(V_on), cmath.phase(V_on), V_on.real, V_on.imag, V_on],
        }

    df = pd.DataFrame(data_dict)
    # Adicionar o DataFrame à planilha do Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):  # começando na linha 1
        for c_idx, value in enumerate(row, 1):
            value = str(value[0]) if isinstance(value, np.ndarray) and isinstance(value[0], np.complex128) else value
            value = value[0] if isinstance(value, np.ndarray) else value
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Salvar a planilha do Excel
    wb.save(excel_filename)
    del wb



def round_complex_array(z_array, places=12):
    real_part = np.round(z_array.real, places)
    imag_part = np.round(z_array.imag, places)
    return real_part + imag_part*1j



import openpyxl
from openpyxl.styles import Border, Side

import openpyxl
from openpyxl.styles import Border, Side


def highlight_cells_above_threshold(filename, sheetnames, threshold_value):
    """
    Adiciona bordas vermelhas espessas às células em um arquivo Excel
    que sejam superiores ao valor de limite fornecido.

    Args:
    - filename (str): O caminho para o arquivo Excel.
    - sheetnames (list): Uma lista com os nomes das planilhas que serão verificadas.
    - threshold_value (float): O valor de limite para destacar as células.
    """

    wb = openpyxl.load_workbook(filename)

    thick_red_border = Border(left=Side(color="FF0000", border_style="thick"),
                              right=Side(color="FF0000", border_style="thick"),
                              top=Side(color="FF0000", border_style="thick"),
                              bottom=Side(color="FF0000", border_style="thick"))

    for sheetname in sheetnames:
        ws = wb[sheetname]

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value > threshold_value:
                    cell.border = thick_red_border

    wb.save(filename)

